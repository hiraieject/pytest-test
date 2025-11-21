"""
Excelレポート生成モジュール
pytestの実行結果をExcelファイルに出力
"""
import os
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger("pytest_logger")


class ExcelReporter:
    """Excelレポート生成クラス"""
    
    def __init__(self, output_dir: str = "output"):
        """
        初期化
        
        Args:
            output_dir: 出力ディレクトリのパス
        """
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.excel_file = os.path.join(output_dir, f"test_results_{self.timestamp}.xlsx")
        self.workbook = None
        self.test_results: List[Dict[str, Any]] = []
        
    def initialize_workbook(self):
        """ワークブックの初期化"""
        self.workbook = Workbook()
        # デフォルトシートを削除
        if "Sheet" in self.workbook.sheetnames:
            del self.workbook["Sheet"]
        logger.info(f"Excelワークブックを初期化しました: {self.excel_file}")
        
    def add_test_result(self, test_name: str, status: str, duration: float, 
                       error_message: str = "", category: str = "General"):
        """
        テスト結果を追加
        
        Args:
            test_name: テスト名
            status: テスト結果 (passed/failed/skipped)
            duration: 実行時間（秒）
            error_message: エラーメッセージ（失敗時）
            category: テストのカテゴリ
        """
        result = {
            "test_name": test_name,
            "status": status,
            "duration": duration,
            "error_message": error_message,
            "category": category,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        self.test_results.append(result)
        logger.debug(f"テスト結果を追加: {test_name} - {status}")
        
    def create_summary_sheet(self):
        """サマリーシートの作成"""
        ws = self.workbook.create_sheet("Summary", 0)
        
        # ヘッダースタイル
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # タイトル
        ws["A1"] = "テスト実行サマリー"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:D1")
        
        # 実行情報
        ws["A3"] = "実行日時:"
        ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 統計情報のヘッダー
        headers = ["項目", "件数", "割合(%)", "備考"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        
        # 統計データの計算
        total = len(self.test_results)
        passed = sum(1 for r in self.test_results if r["status"] == "passed")
        failed = sum(1 for r in self.test_results if r["status"] == "failed")
        skipped = sum(1 for r in self.test_results if r["status"] == "skipped")
        
        stats = [
            ("総テスト数", total, 100.0, ""),
            ("成功", passed, (passed/total*100 if total > 0 else 0), ""),
            ("失敗", failed, (failed/total*100 if total > 0 else 0), ""),
            ("スキップ", skipped, (skipped/total*100 if total > 0 else 0), "")
        ]
        
        # 統計データの書き込み
        for row_idx, (item, count, percentage, note) in enumerate(stats, start=6):
            ws.cell(row=row_idx, column=1, value=item).border = border
            ws.cell(row=row_idx, column=2, value=count).border = border
            ws.cell(row=row_idx, column=3, value=f"{percentage:.1f}").border = border
            ws.cell(row=row_idx, column=4, value=note).border = border
            
            # 結果に応じた色付け
            if item == "成功" and count > 0:
                ws.cell(row=row_idx, column=2).fill = PatternFill(start_color="C6EFCE", 
                                                                   end_color="C6EFCE", 
                                                                   fill_type="solid")
            elif item == "失敗" and count > 0:
                ws.cell(row=row_idx, column=2).fill = PatternFill(start_color="FFC7CE", 
                                                                   end_color="FFC7CE", 
                                                                   fill_type="solid")
        
        # カラム幅の調整
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 30
        
        logger.info("サマリーシートを作成しました")
        
    def create_detail_sheet(self, category: str = None):
        """
        詳細結果シートの作成
        
        Args:
            category: フィルタするカテゴリ（Noneの場合は全件）
        """
        sheet_name = category if category else "All Tests"
        ws = self.workbook.create_sheet(sheet_name)
        
        # ヘッダースタイル
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ヘッダー
        headers = ["No.", "テスト名", "カテゴリ", "結果", "実行時間(秒)", "実行日時", "エラーメッセージ"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        
        # データのフィルタリング
        filtered_results = [r for r in self.test_results 
                          if category is None or r["category"] == category]
        
        # データの書き込み
        for row_idx, result in enumerate(filtered_results, start=2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1).border = border
            ws.cell(row=row_idx, column=2, value=result["test_name"]).border = border
            ws.cell(row=row_idx, column=3, value=result["category"]).border = border
            
            # 結果セルの設定
            status_cell = ws.cell(row=row_idx, column=4, value=result["status"])
            status_cell.border = border
            if result["status"] == "passed":
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", 
                                              fill_type="solid")
            elif result["status"] == "failed":
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", 
                                              fill_type="solid")
            elif result["status"] == "skipped":
                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", 
                                              fill_type="solid")
            
            ws.cell(row=row_idx, column=5, value=f"{result['duration']:.3f}").border = border
            ws.cell(row=row_idx, column=6, value=result["timestamp"]).border = border
            ws.cell(row=row_idx, column=7, value=result["error_message"]).border = border
        
        # カラム幅の調整
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 50
        
        logger.info(f"詳細シートを作成しました: {sheet_name}")
        
    def save(self):
        """Excelファイルの保存"""
        if self.workbook:
            # カテゴリ別のシート作成
            categories = set(r["category"] for r in self.test_results)
            for category in sorted(categories):
                self.create_detail_sheet(category)
            
            # 全体の詳細シート作成
            self.create_detail_sheet()
            
            # サマリーシート作成（最初に表示されるように）
            self.create_summary_sheet()
            
            # 保存
            self.workbook.save(self.excel_file)
            logger.info(f"Excelファイルを保存しました: {self.excel_file}")
            return self.excel_file
        return None
