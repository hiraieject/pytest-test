"""
テンプレートベースExcel管理モジュール
テンプレートファイルをコピーして結果を記録する
"""
import os
import shutil
from datetime import datetime
from typing import Optional
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import logging

logger = logging.getLogger("pytest_logger")

# テンプレートファイルの構造定数
HEADER_ROW = 3
FIRST_DATA_ROW = 4


class ExcelManager:
    """テンプレートベースExcel管理クラス"""
    
    def __init__(self, template_path: str, output_path: str, 
                 sheet_name: str, rom_version: str, tester_name: str):
        """
        初期化
        
        Args:
            template_path: テンプレートファイルのパス
            output_path: 出力ファイルのパス
            sheet_name: 出力シート名
            rom_version: ターゲットROMバージョン
            tester_name: テスト実施者氏名
        """
        self.template_path = template_path
        self.output_path = output_path
        self.sheet_name = sheet_name
        self.rom_version = rom_version
        self.tester_name = tester_name
        self.workbook = None
        self.worksheet = None
    
    def prepare_output_file(self):
        """出力ファイルの準備（テンプレートのコピー）"""
        # 出力ディレクトリの存在確認
        output_dir = os.path.dirname(self.output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
        
        # 出力ファイルが存在しない場合、テンプレートをコピー
        if not os.path.exists(self.output_path):
            if not os.path.exists(self.template_path):
                raise FileNotFoundError(f"テンプレートファイルが見つかりません: {self.template_path}")
            
            shutil.copy2(self.template_path, self.output_path)
            logger.info(f"テンプレートファイルをコピーしました: {self.template_path} -> {self.output_path}")
        else:
            logger.info(f"既存の出力ファイルを使用します: {self.output_path}")
    
    def open_workbook(self):
        """ワークブックを開く"""
        if not os.path.exists(self.output_path):
            raise FileNotFoundError(f"出力ファイルが見つかりません: {self.output_path}")
        
        self.workbook = load_workbook(self.output_path)
        
        # シートの取得または作成
        if self.sheet_name in self.workbook.sheetnames:
            self.worksheet = self.workbook[self.sheet_name]
            logger.info(f"既存のシートを使用します: {self.sheet_name}")
        else:
            # シートが存在しない場合はエラー（テンプレートに含まれているべき）
            raise ValueError(f"シート '{self.sheet_name}' がテンプレートに存在しません")
    
    def write_test_info(self):
        """テスト諸情報の書き込み"""
        if not self.worksheet:
            raise RuntimeError("ワークブックが開かれていません")
        
        # 1行目にROMバージョンを記載（B1セルに書き込む）
        # A1には "ターゲットROMバージョン:" というラベルが既に記載されていると仮定
        self.worksheet['B1'] = self.rom_version
        logger.info(f"ROMバージョンを記載しました: {self.rom_version}")
    
    def find_test_row(self, test_id: str) -> Optional[int]:
        """
        テスト番号からテスト行を検索
        
        Args:
            test_id: テスト番号
            
        Returns:
            行番号（見つからない場合はNone）
        """
        if not self.worksheet:
            return None
        
        # FIRST_DATA_ROW以降でテスト番号列（A列）を検索
        for row in range(FIRST_DATA_ROW, self.worksheet.max_row + 1):
            cell_value = self.worksheet.cell(row=row, column=1).value
            if cell_value and str(cell_value).strip() == str(test_id).strip():
                return row
        
        return None
    
    def _format_test_date(self) -> str:
        """
        テスト実施日付をフォーマット
        
        Returns:
            フォーマットされた日付（月/日形式）
        """
        now = datetime.now()
        # 月/日形式で、先頭のゼロを削除
        return f"{now.month}/{now.day}"
    
    def write_test_result(self, test_id: str, status: str, error_info: str = ""):
        """
        テスト結果の書き込み
        
        Args:
            test_id: テスト番号
            status: テスト結果 (passed/failed/skipped)
            error_info: エラー情報（NGの場合）
        """
        if not self.worksheet:
            raise RuntimeError("ワークブックが開かれていません")
        
        row = self.find_test_row(test_id)
        if row is None:
            logger.warning(f"テスト番号 '{test_id}' がテンプレートに見つかりません。スキップします。")
            return
        
        # テスト実施日付を記入（F列）
        test_date = self._format_test_date()
        self.worksheet.cell(row=row, column=6, value=test_date)
        
        # テスト結果を記入（G列）
        result_text = "OK" if status == "passed" else "NG" if status == "failed" else "SKIP"
        result_cell = self.worksheet.cell(row=row, column=7, value=result_text)
        
        # 結果に応じた色付け
        if status == "passed":
            result_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif status == "failed":
            result_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # テスト結果補足を記入（H列、NGの場合のみ）
        if status == "failed" and error_info:
            self.worksheet.cell(row=row, column=8, value=error_info)
        
        logger.info(f"テスト結果を記録しました: {test_id} -> {result_text} (行{row})")
    
    def save(self):
        """ワークブックの保存"""
        if self.workbook:
            self.workbook.save(self.output_path)
            logger.info(f"テスト結果を保存しました: {self.output_path}")
    
    def close(self):
        """ワークブックを閉じる"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.worksheet = None
